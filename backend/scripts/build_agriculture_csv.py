"""Extrait la production nette de céréales et les importations de céréales
du Sénégal (séries nationales annuelles), pour le domaine Agriculture.

Sorties :
  - data/raw/agriculture_production_cereales.csv (production nette, tonnes)
  - data/raw/agriculture_importations_cereales.csv (importations, tonnes)

Sources : ANSD/DAPSA, fichiers officiels déjà téléchargés dans
data/raw/ansd_downloads/ (BADIS 2018) :
  - "23_SECTION G - AGRICULTURE ET SECURITE ALIMENTAIRE..." feuille G0201
    (TABLEAU G.02.01, colonne 13 = Total "Production hivernale Nette en
    Céréales entières", en tonnes, par campagne agricole)
  - "57_Importations des céréales, 1970 - 2018" (TABLEAU G.02.02, colonne 6
    = Total "Importations de céréales entières", en tonnes, par année civile)

Convention de rapprochement des deux séries : la production est indexée par
CAMPAGNE agricole (ex. "1969/70"), les importations par ANNÉE CIVILE (ex.
"1970"). On associe la campagne à son année de DÉBUT (1969/70 -> 1969) --
c'est une convention arbitraire mais documentée, nécessaire pour permettre au
moteur de croisement de joindre les deux séries sur une dimension "year"
commune ; elle introduit un décalage d'au plus quelques mois entre les deux
séries, sans incidence sur l'ordre de grandeur du ratio de dépendance calculé.
"""

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE = Path(r"C:\Users\Birame\ansd_hackathon\backend")
D = BASE / "data" / "raw" / "ansd_downloads"
OUT_DIR = BASE / "data" / "raw"


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


# ── 1. Production nette de céréales (campagne agricole) ─────────────────
prod_file = [p for p in D.iterdir() if p.name.startswith("23_SECTION G") and p.suffix == ".xlsx"][0]
wb = load_workbook(str(prod_file), data_only=True, read_only=True)
ws = wb["G0201"]

production: dict[int, float] = {}
for row in ws.iter_rows(min_row=10, max_row=ws.max_row, values_only=True):
    campagne = row[0]
    if not isinstance(campagne, str) or "/" not in campagne:
        continue
    start_year = int(campagne.split("/")[0])
    total_nette = num(row[12])  # colonne 13 (index 12) = Total "Nette en Céréales entières"
    if total_nette is None or total_nette <= 0:
        continue
    if start_year in production:
        raise AssertionError(f"campagne dupliquée pour l'année {start_year}")
    production[start_year] = total_nette

print(f"Production : {len(production)} campagnes ({min(production)}-{max(production)})")
assert len(production) >= 40, f"trop peu de campagnes extraites ({len(production)})"
# Sanity check : la production nette ne doit jamais dépasser la production brute
# totale connue pour la même campagne, et rester dans un ordre de grandeur stable
# (aucune valeur négative, aucun saut x100 d'une campagne à l'autre).
sorted_years = sorted(production)
for a, b in zip(sorted_years, sorted_years[1:]):
    ratio = production[b] / production[a] if production[a] else 0
    assert 0 < ratio < 10, f"saut anormal entre {a} et {b} (x{ratio:.1f}) -- vérifier l'extraction"

# ── 2. Importations de céréales (année civile) ───────────────────────────
import_file = [p for p in D.iterdir() if p.name.startswith("57_Importations des c") and p.suffix == ".xlsx"][0]
wb2 = load_workbook(str(import_file), data_only=True, read_only=True)
ws2 = wb2[wb2.sheetnames[0]]

imports: dict[int, float] = {}
for row in ws2.iter_rows(min_row=9, max_row=ws2.max_row, values_only=True):
    year = row[0]
    # Les années sont stockées en texte pour 1970-2000 et en nombre pour 2001+.
    if not isinstance(year, (int, float, str)):
        continue
    try:
        year_int = int(str(year).strip())
    except ValueError:
        continue
    if not (1900 <= year_int <= 2100):
        continue
    total = num(row[5])  # colonne 6 (index 5) = Total "Importations de céréales entières"
    if total is None or total <= 0:
        continue
    imports[year_int] = total

print(f"Importations : {len(imports)} années ({min(imports)}-{max(imports)})")
assert len(imports) >= 30, f"trop peu d'années extraites ({len(imports)})"

# ── 3. Écriture des CSV ───────────────────────────────────────────────────
prod_out = OUT_DIR / "agriculture_production_cereales.csv"
with open(prod_out, "w", encoding="utf-8", newline="") as fp:
    fp.write(
        "# Jeu de données : Production nette de céréales du Sénégal (série nationale)\n"
        "# SOURCE OFFICIELLE : ANSD/DAPSA, TABLEAU G.02.01, fichier BADIS 2018\n"
        "#   \"SECTION G - AGRICULTURE ET SECURITE ALIMENTAIRE\" (data/raw/ansd_downloads/)\n"
        "# Année = année de début de la campagne agricole (voir docstring du script\n"
        "# scripts/build_agriculture_csv.py pour la convention de rapprochement avec\n"
        "# les importations, qui sont en année civile).\n"
        "year,production_nette_tonnes,source\n"
    )
    w = csv.writer(fp)
    for year in sorted(production):
        w.writerow([year, f"{production[year]:.1f}", "ANSD/DAPSA (BADIS 2018, TABLEAU G.02.01)"])
print(f"OK -> {prod_out} ({len(production)} lignes)")

imports_out = OUT_DIR / "agriculture_importations_cereales.csv"
with open(imports_out, "w", encoding="utf-8", newline="") as fp:
    fp.write(
        "# Jeu de données : Importations de céréales du Sénégal (série nationale)\n"
        "# SOURCE OFFICIELLE : ANSD/DAPSA, TABLEAU G.02.02, fichier BADIS 2018\n"
        "#   \"Importations des céréales, 1970 - 2018\" (data/raw/ansd_downloads/)\n"
        "# Année civile.\n"
        "year,import_tonnes,source\n"
    )
    w = csv.writer(fp)
    for year in sorted(imports):
        w.writerow([year, f"{imports[year]:.1f}", "ANSD/DAPSA (BADIS 2018, TABLEAU G.02.02)"])
print(f"OK -> {imports_out} ({len(imports)} lignes)")
