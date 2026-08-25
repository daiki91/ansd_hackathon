"""Extrait l'inflation officielle ANSD depuis l'IHPC base 2023 (indice Ensemble,
moyennes annuelles). Régénère indicateurs_nationaux.csv avec :
  - Croissance PIB (DG Trésor citant sources officielles) : conservée
  - Inflation IHPC calculée depuis l'indice officiel ANSD (remplace les valeurs)
"""
import sys
import csv
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = Path(r'C:\Users\Birame\ansd_hackathon\backend')
D = BASE / 'data' / 'raw' / 'ansd_downloads'
f = [p for p in D.iterdir() if 'IHPC_base2023' in p.name][0]
wb = load_workbook(str(f), read_only=True, data_only=True)

# ── Trouver la ligne ENSEMBLE dans DIV ──
sh = wb['DIV']
rows = list(sh.iter_rows(values_only=True))
ens_row = None
hdr_i = None
for i, row in enumerate(rows):
    cells = [str(c).strip().upper() if c else '' for c in row[:5]]
    if any(c in ('ENSEMBLE', 'GLOBAL') for c in cells):
        ens_row = row
        print(f'ligne indice général trouvée à la ligne {i+1}')
        break

if ens_row is None:
    # fallback : ligne avec pondération ~100%
    for row in rows[6:30]:
        if isinstance(row[3], float) and 0.99 < float(row[3]) < 1.01:
            ens_row = row
            print('fallback pondération=1 utilisée:', str(row[2])[:40])
            break

assert ens_row is not None, 'ligne ENSEMBLE introuvable'
print('Ligne indice général:', [str(x)[:20] for x in ens_row[:4]])

# colonnes mois
months = {}
for j in range(4, len(ens_row)):
    pass

# Les dates sont sur la ligne d'en-tête ; retrouver via rows[hdr_i] si dispo sinon DIV row 7
date_row = None
for r in rows[:10]:
    vals = list(r)
    hits = sum(1 for v in vals if v is not None and str(v)[:2] in ('19', '20') and len(str(v)) >= 8)
    if hits > 50:
        date_row = vals
        break
assert date_row is not None, 'ligne de dates introuvable'

series = {}   # year -> [monthly index...]
for j in range(len(date_row)):
    d = date_row[j]
    if d is None:
        continue
    try:
        yr = int(str(d)[:4])
    except ValueError:
        continue
    v = ens_row[j]
    if isinstance(v, (int, float)):
        series.setdefault(yr, []).append(float(v))

annual_idx = {y: sum(vals) / len(vals) for y, vals in sorted(series.items()) if len(vals) == 12}
print(f'Indices annuels complets: {min(annual_idx)}..{max(annual_idx)}')

inflation = {}
years_sorted = sorted(annual_idx)
for prev, cur in zip(years_sorted, years_sorted[1:]):
    inflation[cur] = round((annual_idx[cur] / annual_idx[prev] - 1) * 100, 1)

for y in sorted(inflation)[-6:]:
    print(f'  Inflation moyenne {y}: {inflation[y]} %')

# ── Régénérer indicateurs_nationaux.csv ──
out = BASE / 'data' / 'raw' / 'indicateurs_nationaux.csv'
with open(out, 'w', encoding='utf-8', newline='') as fp:
    fp.write("""# Jeu de données : Indicateurs nationaux dynamiques -- Économie & Population
# SOURCES OFFICIELLES :
#  - Inflation (IHPC, moyennes annuelles calculées depuis l'indice Ensemble base 2023) :
#      ANSD, "Série IHPC base 2023" téléchargé depuis https://www.ansd.sn/toutes-les-publications
#  - Croissance du PIB 2020-2024 : Direction générale du Trésor (France) citant
#    les statistiques officielles sénégalaises :
#    https://www.tresor.economie.gouv.fr/Pays/SN/situation-economique-et-financiere-du-senegal
#  - Indicateurs démographiques : ANSD RGPH-5 (2023).
category,indicator,value,unit,year,source
""")
    w = csv.writer(fp)
    gdp_rows = [
        (2020, 1.5), (2021, 6.5), (2022, 4.0), (2023, 4.1), (2024, 4.5),
    ]
    for y, v in gdp_rows:
        w.writerow(['Économie', 'Croissance du PIB', v, '%', y, 'DG Trésor (France) citant sources officielles'])
    for y in sorted(inflation):
        w.writerow(['Économie', 'Inflation (IHPC)', inflation[y], '%', y,
                    'ANSD IHPC base 2023 (calcul: moyenne annuelle des indices)'])
    w.writerow(['Économie', 'Dette publique', 80.0, '% du PIB', 2023, 'DG Trésor (France) citant sources officielles'])
    w.writerow(['Population', 'Taux de croissance démographique annuel', 2.9, '%/an', 2023, 'ANSD (RGPH-5 -- période 2013-2023)'])
    w.writerow(['Population', "Taux d'urbanisation", 54.7, '%', 2023, 'ANSD (RGPH-5)'])
    w.writerow(['Population', 'Espérance de vie à la naissance', 68.9, 'années', 2023, 'ANSD (RGPH-5)'])

print(f'OK -> {out}')
