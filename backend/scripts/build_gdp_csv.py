"""Extrait le PIB régional officiel ANSD (comptes régionaux 2020-2023).
Sortie : data/raw/regional_gdp.csv
  - PIB en volume et en valeur par région/année (feuille Sénégal)
  - Structure sectorielle de la VA par région (feuilles régionales, 2023)
"""
import sys
import csv
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = Path(r'C:\Users\Birame\ansd_hackathon\backend')
D = BASE / 'data' / 'raw' / 'ansd_downloads'
f = [p for p in D.iterdir() if 'comptes r' in p.name.lower()][0]
wb = load_workbook(str(f), read_only=True)

YEARS = [2020, 2021, 2022, 2023]
REGION_SHEET = {'Dakar': 'Dakar', 'Diourbel': 'Diourbel', 'Fatick': 'Fatick',
                'Kaffrine': 'Kaffrine', 'Kaolack': 'Kaolack', 'Kédougou': 'Kédougou',
                'Kolda': 'Kolda', 'Louga': 'Louga', 'Matam': 'Matam',
                'Saint-Louis': 'Saint-Louis', 'Sédhiou': 'Sédhiou',
                'Tambacounda': 'Tamba', 'Thiès': 'Thiès', 'Ziguinchor': 'Ziguinchor'}

def num(v):
    return float(v) if isinstance(v, (int, float)) else None

# ── 1. Feuille nationale : PIB volume + valeur par région ──
sh = wb['Sénégal']
grid = [[c for c in row] for row in sh.iter_rows(values_only=True)]

def read_block(start_label):
    """Lit les ~14 régions suivant le libellé de bloc, s'arrête à 'Total'."""
    start = None
    for r in range(len(grid)):
        if grid[r][0] and start_label.lower() in str(grid[r][0]).lower():
            start = r + 1
            break
    out = {}
    if start is None:
        return out
    for r in range(start, min(start + 16, len(grid))):
        name = grid[r][0]
        if not name:
            continue
        label = str(name).strip()
        if label.lower() == 'total':
            break
        vals = {YEARS[i]: num(grid[r][1 + i]) for i in range(4)}
        if all(vals.values()):
            out[label] = vals
    return out

pib_vol = read_block('PIB en volume')
pib_val = read_block('PIB en valeur')

print(f'PIB volume régions: {len(pib_vol)} | PIB valeur: {len(pib_val)}')
assert len(pib_vol) == 14, f'attendu 14 régions volume, trouvé {len(pib_vol)}'
assert len(pib_val) == 14, f'attendu 14 régions valeur, trouvé {len(pib_val)}'

# ── 2. Feuilles régionales : structure sectorielle VA ──
sectors = {}   # region -> {year: (prim, sec, tert)}
for region, sheet in REGION_SHEET.items():
    if sheet not in wb.sheetnames:
        continue
    s = wb[sheet]
    g = [[c for c in row] for row in s.iter_rows(min_row=10, max_row=17, max_col=6, values_only=True)]
    # g[1]=VA totale, g[2]=primaire, g[3]=secondaire, g[4]=tertiaire ; colonnes 1..4 = 2020..2023
    try:
        d = {}
        for i, yr in enumerate(YEARS):
            prim, sec, tert = num(g[2][1 + i]), num(g[3][1 + i]), num(g[4][1 + i])
            va_tot = num(g[1][1 + i])
            if None not in (prim, sec, tert) and va_tot:
                d[yr] = (round(prim * 100 / va_tot, 1), round(sec * 100 / va_tot, 1), round(tert * 100 / va_tot, 1))
        sectors[region] = d
    except Exception as e:
        print(f'  secteurs {region}: skip ({e})')

print(f'Structure sectorielle: {len(sectors)} régions')

# ── 3. Vérification : somme régionale vs Total national ──
tot_2023 = sum(pib_vol[r][2023] for r in pib_vol)
print(f"Somme volume 2023: {tot_2023:.0f} Mds (attendu ~15428)")
assert abs(tot_2023 - 15427.9) < 5, 'somme incohérente avec le Total ANSD'

# ── 4. Écriture CSV ──
out = BASE / 'data' / 'raw' / 'regional_gdp.csv'
with open(out, 'w', encoding='utf-8', newline='') as fp:
    fp.write("""# Produit Intérieur Brut Régional du Sénégal
# SOURCE OFFICIELLE EXACTE : ANSD, "Résultats des comptes économiques régionaux 2020-2023"
#   téléchargé depuis https://www.ansd.sn/toutes-les-publications
# PIB en volume et en valeur (milliards FCFA). Part sectorielle = % de la Valeur Ajoutée totale.
region,year,pib_volume_mds,pib_valeur_mds,part_primaire_pct,part_secondaire_pct,part_tertiaire_pct
""")
    w = csv.writer(fp)
    n = 0
    for region in sorted(pib_vol):
        for yr in YEARS:
            vol = pib_vol[region].get(yr, '')
            val = pib_val.get(region, {}).get(yr, '')
            sec = sectors.get(region, {}).get(yr, ('', '', ''))
            w.writerow([region, yr,
                        f'{vol:.1f}' if vol else '',
                        f'{val:.1f}' if val else '',
                        sec[0], sec[1], sec[2]])
            n += 1
print(f'OK -> {out} ({n} lignes)')
