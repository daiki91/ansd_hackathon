import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

D = Path(r'C:\Users\Birame\ansd_hackathon\backend\data\raw\ansd_downloads')

f = [p for p in D.iterdir() if 'hopitaux' in p.name.lower() or ('volution' in p.name and 'sant' in p.name.lower())][0]
print('FILE:', f.name, f.stat().st_size // 1024, 'Ko\n')

wb = load_workbook(str(f), data_only=True)
print('SHEETS:', wb.sheetnames, '\n')

for sn in wb.sheetnames:
    sh = wb[sn]
    print(f'===== SHEET "{sn}" ({sh.max_row}x{sh.max_column}) =====')
    for r in range(1, min(sh.max_row, 60) + 1):
        vals = []
        for c in range(1, min(sh.max_column, 16) + 1):
            v = sh.cell(r, c).value
            if isinstance(v, float) and v == int(v):
                v = int(v)
            vals.append(str(v)[:18] if v is not None else '')
        line = ' | '.join(vals).rstrip(' |')
        if line.strip():
            print(f'  {r:3d}: {line[:200]}')
    print()
