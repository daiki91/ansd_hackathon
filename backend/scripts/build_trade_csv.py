"""Extrait le commerce extérieur réel ANSD (imports/exports par partenaire,
séries mensuelles janv 2010 - févr 2026). Agrège en annuel + parts partenaires.
Sortie : data/raw/commerce_exterieur.csv (compatible modèle TradeFlow).
"""
import sys
import csv
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = Path(r'C:\Users\Birame\ansd_hackathon\backend')
D = BASE / 'data' / 'raw' / 'ansd_downloads'
f = [p for p in D.iterdir() if 'imports_exports_janv2010-fevr2026' in p.name.lower()][0]
wb = load_workbook(str(f), read_only=True, data_only=True)


def clean_country(raw):
    s = str(raw).strip()
    if ':' in s:
        s = s.split(':', 1)[1].strip()
    return s.title().replace('Pays Bas', 'Pays-Bas').replace('Royaume Uni', 'Royaume-Uni')


def parse_sheet(sheet_name, flow_label):
    sh = wb[sheet_name]
    rows = list(sh.iter_rows(values_only=True))
    # ligne d'en-tête : première cellule 'PARTENAIRE', colonnes = mois
    hdr_i = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip().upper() == 'PARTENAIRE':
            hdr_i = i
            break
    assert hdr_i is not None, f'en-tête introuvable dans {sheet_name}'
    months = rows[hdr_i][1:]
    cols = []   # (col_idx, year)
    for j, m in enumerate(months):
        if m is None:
            continue
        try:
            yr = int(str(m)[:4])
            cols.append((j + 1, yr))
        except ValueError:
            continue
    years = sorted(set(y for _, y in cols))
    print(f'{sheet_name}: {len(cols)} colonnes mensuelles, {years[0]}..{years[-1]}')

    per_year_total = {y: 0.0 for y in years}
    per_year_country = {}   # country -> {year: value}
    for row in rows[hdr_i + 1:]:
        name = row[0]
        if name is None or str(name).strip() == '':
            continue
        if str(name).strip().lower() in ('total', 'totaux', 'autres', 'autre'):
            continue   # lignes agrégées du fichier source : on recalcule nous-mêmes
        country = clean_country(name)
        agg = {}
        for j, yr in cols:
            v = row[j]
            if isinstance(v, (int, float)):
                agg[yr] = agg.get(yr, 0) + float(v)
        if not agg:
            continue
        per_year_country[country] = agg
        for y, v in agg.items():
            per_year_total[y] += v

    return years, per_year_total, per_year_country


imp_years, imp_tot, imp_cty = parse_sheet('Import_Partenaire ', 'Importation')
exp_years, exp_tot, exp_cty = parse_sheet('Export_Partenaire ', 'Exportation')

rows_out = []
for years, tot, cty, flow in [
    (imp_years, imp_tot, imp_cty, 'Importation'),
    (exp_years, exp_tot, exp_cty, 'Exportation'),
]:
    for y in years:
        tv = tot.get(y, 0)
        if tv <= 0:
            continue
        top = sorted(((c, a.get(y, 0)) for c, a in cty.items()), key=lambda x: -x[1])
        # Top 15 partenaires + part
        for country, v in top[:15]:
            if v <= 0:
                continue
            rows_out.append({
                'year': y,
                'flow_type': flow,
                'country': country,
                'share_pct': round(v * 100 / tv, 2),
                'value_fcfa_billions': round(v / 1e9, 2),
                'source': 'ANSD/DGIT (imports_exports 2010-2026)',
            })
        # Autre (agrégé)
        rest = sum(v for _, v in top[15:] if v > 0)
        if rest > 0:
            rows_out.append({'year': y, 'flow_type': flow, 'country': 'Autres',
                             'share_pct': round(rest * 100 / tv, 2),
                             'value_fcfa_billions': round(rest / 1e9, 2),
                             'source': 'ANSD/DGIT (imports_exports 2010-2026)'})
        # TOTAL national
        rows_out.append({'year': y, 'flow_type': flow, 'country': 'Total Sénégal',
                         'share_pct': '', 'value_fcfa_billions': round(tv / 1e9, 2),
                         'source': 'ANSD/DGIT (imports_exports 2010-2026)'})

# Vérifications
t2025_imp = [r for r in rows_out if r['year'] == 2025 and r['flow_type'] == 'Importation' and r['country'] == 'Total Sénégal']
print(f"\nImportations 2025 : {float(t2025_imp[0]['value_fcfa_billions']):,.0f} Mds FCFA")
top3_2025 = [r['country'] for r in rows_out if r['year'] == 2025 and r['flow_type'] == 'Importation'][:3]
print(f'Top import partners 2025: {top3_2025}')

out = BASE / 'data' / 'raw' / 'commerce_exterieur.csv'
with open(out, 'w', encoding='utf-8', newline='') as fp:
    fp.write("""# Commerce extérieur du Sénégal - exportations et importations par pays
# SOURCE OFFICIELLE EXACTE : ANSD/DGIT, "Imports et exports janvier 2010 - février 2026"
#   téléchargé depuis https://www.ansd.sn/toutes-les-publications
# Valeurs douanières (VALDOUANE), agrégation annuelle des séries mensuelles.
# Top 15 partenaires + "Autres" ; "Total Sénégal" = total national.
year,flow_type,country,share_pct,value_fcfa_billions,source
""")
    w = csv.DictWriter(fp, fieldnames=['year', 'flow_type', 'country', 'share_pct', 'value_fcfa_billions', 'source'])
    w.writerows(rows_out)

print(f'OK -> {out} ({len(rows_out)} lignes)')
