"""Explore la structure des fichiers XLSX ANSD à exploiter."""
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

D = Path(r'C:\Users\Birame\ansd_hackathon\backend\data\raw\ansd_downloads')


def peek(fname_pattern, max_rows=18, max_cols=14, sheets_limit=3):
    files = [p for p in D.iterdir() if fname_pattern.lower() in p.name.lower()]
    if not files:
        print(f'!! introuvable: {fname_pattern}')
        return
    f = files[0]
    print(f'######## {f.name} ({f.stat().st_size//1024} Ko)')
    try:
        wb = load_workbook(str(f), data_only=True, read_only=True)
    except Exception as e:
        print('  ERR', e)
        return
    print('SHEETS:', wb.sheetnames[:10])
    for sn in wb.sheetnames[:sheets_limit]:
        sh = wb[sn]
        print(f'--- "{sn}" ---')
        for i, row in enumerate(sh.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True)):
            vals = []
            for v in row:
                if isinstance(v, float) and v == int(v):
                    v = int(v)
                vals.append(str(v)[:20] if v is not None else '')
            line = ' | '.join(vals).rstrip(' |')
            if line.strip():
                print(f'  {i+1:3d}: {line[:190]}')
    wb.close()
    print()


peek('comptes r', sheets_limit=3)
peek('imports_exports_janv2010-fevr2026', sheets_limit=2, max_cols=8)
