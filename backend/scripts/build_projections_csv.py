"""Parse le fichier officiel ANSD Projections-demographiques_2023-2050.xlsx.
Extrait régions (14) ET départements (45), par année, colonne ENSEMBLE.
Génère :
  - data/raw/projections_regions_2023_2050.csv
  - data/raw/projections_departements_2023_2050.csv
"""
import sys
import csv
import re
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = Path(r'C:\Users\Birame\ansd_hackathon\backend')
f = [p for p in (BASE / 'data' / 'raw' / 'ansd_downloads').iterdir() if 'Projections-demographiques_2023-2050' in p.name][0]
wb = load_workbook(str(f), data_only=True)
sh = wb[wb.sheetnames[0]]

# ── 1. Mapper les années aux groupes de colonnes ──
# Ligne 1 : "PROJECTION DE LA", "RGPH-5 2023", ..., "2024", ..., chaque groupe = 3 colonnes (H,F,E)
year_cols = {}   # year -> index colonne ENSEMBLE (0-based)
current_year = None
c = 2
max_col = sh.max_column
while c < max_col + 2:
    v = sh.cell(1, c).value
    if v is not None:
        s = str(v).strip()
        m = re.match(r'(?:RGPH-?5?\s*)?(\d{4})', s)
        if m or s.isdigit():
            yr = int(m.group(1)) if m else int(s)
            current_year = yr
            year_cols[yr] = c + 2   # ENSEMBLE = H+2
    c += 1

years = sorted(year_cols)
print(f'Années détectées ({len(years)}): {years[:6]} ... {years[-3:]}')

def clean_name(s):
    s = str(s).replace('REGION ', '').replace('DEPARTEMENT DE ', '').replace('DEPARTEMENT ', '').strip()
    return re.sub(r'\s+', ' ', s)

REGION_NAMES = {
    'DAKAR': 'Dakar', 'DIOURBEL': 'Diourbel', 'FATICK': 'Fatick',
    'KAFFRINE': 'Kaffrine', 'KAOLACK': 'Kaolack', 'KEDOUGOU': 'Kédougou',
    'KOLDA': 'Kolda', 'LOUGA': 'Louga', 'MATAM': 'Matam',
    'SAINT-LOUIS': 'Saint-Louis', 'SEDHIOU': 'Sédhiou',
    'TAMBACOUNDA': 'Tambacounda', 'THIES': 'Thiès', 'ZIGUINCHOR': 'Ziguinchor',
}

DEP_NAMES = {
    'DAKAR': 'Dakar', 'GUEDIAWAYE': 'Guédiawaye', 'PIKINE': 'Pikine',
    'RUFISQUE': 'Rufisque', 'KEUR MASSAR': 'Keur Massar', 'BAMBEY': 'Bambey',
    'DIOURBEL': 'Diourbel', 'MBACKE': 'Mbacké', 'FATICK': 'Fatick',
    'FOUNDIOUGNE': 'Foundiougne', 'GOSSAS': 'Gossas', 'KAFFRINE': 'Kaffrine',
    'MBIRKILANE': 'Mbirikilane', 'KOUNGHEUL': 'Koungheul',
    'MALEM HODDAR': 'Malem Hodar', 'KAOLACK': 'Kaolack', 'NIORO': 'Nioro du Rip',
    'GUINGUINEO': 'Guinguinéo', 'KEDOUGOU': 'Kédougou', 'SALEMATA': 'Salémata',
    'SARAYA': 'Saraya', 'KOLDA': 'Kolda', 'VELINGARA': 'Vélingara',
    'MEDINA YORO FOULAH': 'Médina Yoro Foulah', 'KEBEMER': 'Kébémer',
    'LINGUERE': 'Linguère', 'LOUGA': 'Louga', 'MATAM': 'Matam', 'KANEL': 'Kanel',
    'RANEROU': 'Ranérou', 'DAGANA': 'Dagana', 'PODOR': 'Podor',
    'SAINT-LOUIS': 'Saint-Louis', 'SEDHIOU': 'Sédhiou', 'BOUNKILING': 'Bounkiling',
    'GOUDOMP': 'Goudomp', 'BAKEL': 'Bakel', 'TAMBACOUNDA': 'Tambacounda',
    'GOUDIRY': 'Goudiry', 'KOUMPENTOUM': 'Koumpentoum', "M'BOUR": 'Mbour',
    'THIES': 'Thiès', 'TIVAOUANE': 'Tivaouane', 'BIGNONA': 'Bignona',
    'OUSSOUYE': 'Oussouye', 'ZIGUINCHOR': 'Ziguinchor',
}

reg_rows, dep_rows = [], []
for r in range(3, sh.max_row + 1):
    a = sh.cell(r, 1).value
    if a is None:
        continue
    label = str(a).strip()
    if label.upper().startswith('REGION'):
        name = REGION_NAMES.get(clean_name(label).upper(), clean_name(label).title())
        vals = {}
        for yr, colE in year_cols.items():
            v = sh.cell(r, colE).value
            if isinstance(v, (int, float)):
                vals[yr] = int(v)
        reg_rows.append((name, vals))
    elif label.upper().startswith('DEPARTEMENT'):
        name = DEP_NAMES.get(clean_name(label).upper(), clean_name(label).title())
        vals = {}
        for yr, colE in year_cols.items():
            v = sh.cell(r, colE).value
            if isinstance(v, (int, float)):
                vals[yr] = int(v)
        dep_rows.append((name, vals))

print(f'Régions extraites: {len(reg_rows)} | Départements extraits: {len(dep_rows)}')

# ── 2. Vérifications de cohérence ──
names = [n for n, _ in reg_rows]
print('Régions:', names)
assert len(reg_rows) == 14, f'attendu 14 régions, trouvé {len(reg_rows)}'
assert len(dep_rows) >= 45, f'attendu >=45 départements, trouvé {len(dep_rows)}'

# Somme des régions ≈ total national pour chaque année
for yr in years[:3]:
    tot = sum(vals.get(yr, 0) for _, vals in reg_rows)
    print(f'Somme régions {yr}: {tot:,}')

# Cohérence interne : somme départements région = valeur région (échantillon Dakar)
def check_region(region_name):
    rv = dict([x for x in reg_rows if x[0] == region_name][0][1])
    # départements de la région = ceux qui apparaissent après la région dans la feuille

# ── 3. Écriture des CSV ──
def write_csv(path, rows, label):
    with open(path, 'w', encoding='utf-8', newline='') as fp:
        fp.write(f"""# Projections démographiques ANSD ({label}) — base RGPH-5 2023
# SOURCE OFFICIELLE EXACTE : ANSD, "Projections démographiques du Sénégal 2023-2050"
#   téléchargé depuis https://www.ansd.sn/toutes-les-publications
# Valeurs = population totale (ENSEMBLE). 2023 = recensement RGPH-5 ; 2024+ = projections.
# Tant qu'un nouveau recensement n'est pas publié, ces projections font foi.
nom,{','.join(str(y) for y in years)}
""")
        w = csv.writer(fp)
        for name, vals in rows:
            w.writerow([name] + [vals.get(y, '') for y in years])
    print(f'OK -> {path}')

write_csv(BASE / 'data' / 'raw' / 'projections_regions_2023_2050.csv', reg_rows, 'régions')
write_csv(BASE / 'data' / 'raw' / 'projections_departements_2023_2050.csv', dep_rows, 'départements')

# Aperçu Dakar
dk = dict([x for x in reg_rows if x[0] == 'Dakar'][0][1])
print('\nRégions:', [n for n, _ in reg_rows])
print('Départements:', [n for n, _ in dep_rows])
print('\nDakar:', {y: dk.get(y) for y in [2023, 2024, 2025, 2026]})
